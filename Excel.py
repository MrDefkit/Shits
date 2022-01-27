from openpyxl import Workbook, load_workbook

alma = input ("Írd be mi kerüljön bele")

wb = load_workbook('Workbook.xlsx')
ws = wb.active
ws['A2'].value = alma

wb.save('Workbook.xlsx')